import importlib
import json
import os
import shutil
import sys
import unittest
import uuid
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import openpyxl


MODULES_TO_RELOAD = [
    'pnl.config',
    'pnl.utils.logger',
    'pnl.utils.storage',
    'pnl.utils.auth',
    'pnl.utils.validators',
    'pnl.services.pnl_service',
    'pnl.services.excel_service',
    'pnl.routes.auth',
    'pnl.routes.main',
    'pnl.routes.projects',
    'pnl.routes.export',
    'pnl.routes.import_excel',
    'pnl.routes',
    'pnl',
]


def reload_app():
    for name in MODULES_TO_RELOAD:
        if name in sys.modules:
            importlib.reload(sys.modules[name])
    from pnl import create_app
    return create_app()


class ReadOnlyFlowTests(unittest.TestCase):
    def setUp(self):
        temp_root = Path(__file__).resolve().parents[1] / 'test-runtime'
        temp_root.mkdir(parents=True, exist_ok=True)
        self.data_dir = temp_root / f'test-{uuid.uuid4().hex}'
        self.data_dir.mkdir(parents=True, exist_ok=True)

        os.environ['DATA_DIR'] = str(self.data_dir)
        os.environ['SECRET_KEY'] = 'test-secret'
        os.environ['PNL_ENV'] = 'development'
        os.environ.pop('PNL_BOOTSTRAP_ADMIN_PASSWORD', None)

        (self.data_dir / 'projects').mkdir()
        (self.data_dir / 'versions').mkdir()

        self.baseline_data = {
            'project': {'customer': 'Working Copy'},
            'resources': [],
            'rate_card': [{'level': 'L1', 'rate': 10}],
            'role_catalog': [{'group': 'Default', 'roles': ['Engineer']}],
            'business_units': ['EDM', 'AI'],
        }
        (self.data_dir / 'data.json').write_text(json.dumps(self.baseline_data, indent=2), encoding='utf-8')
        (self.data_dir / 'settings.json').write_text(json.dumps({
            'rate_card': [{'level': 'L1', 'rate': 25}],
            'role_catalog': [{'group': 'Delivery', 'roles': ['Architect']}],
        }, indent=2), encoding='utf-8')
        (self.data_dir / 'users.json').write_text(json.dumps({
            'admin': {
                'password': 'unused-in-session-tests',
                'role': 'admin',
                'name': 'Administrator',
                'created_at': '2026-01-01T00:00:00',
            }
        }, indent=2), encoding='utf-8')
        (self.data_dir / 'projects' / 'sample.json').write_text(json.dumps({
            '_meta': {'id': 'sample', 'name': 'Sample'},
            'project': {'customer': 'Saved Project'},
            'resources': [{'role': 'Architect', 'level': 'L1', 'hours': 8}],
            'rate_card': [{'level': 'L1', 'rate': 5}],
            'role_catalog': [{'group': 'Old', 'roles': ['Old Role']}],
        }, indent=2), encoding='utf-8')

        self.app = reload_app()
        self.app.testing = True
        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session['user'] = 'admin'
            session['role'] = 'admin'
            session['name'] = 'Administrator'

    def tearDown(self):
        shutil.rmtree(self.data_dir, ignore_errors=True)

    def _build_excel_bytes(self, build_fn):
        wb = openpyxl.Workbook()
        build_fn(wb)
        bio = BytesIO()
        wb.save(bio)
        wb.close()
        bio.seek(0)
        return bio

    def test_loading_project_does_not_overwrite_working_data(self):
        before = (self.data_dir / 'data.json').read_text(encoding='utf-8')

        response = self.client.get('/api/projects/sample')

        self.assertEqual(response.status_code, 200)
        self.assertEqual((self.data_dir / 'data.json').read_text(encoding='utf-8'), before)

        payload = response.get_json()
        self.assertEqual(payload['project']['customer'], 'Saved Project')
        self.assertEqual(payload['rate_card'][0]['rate'], 25)

    def test_export_does_not_overwrite_working_data(self):
        before = (self.data_dir / 'data.json').read_text(encoding='utf-8')

        response = self.client.post('/api/export', json={
            'project': {'customer': 'Export Project'},
            'resources': [{'role': 'Architect', 'level': 'L1', 'hours': 8}],
            'rate_card': [{'level': 'L1', 'rate': 25}],
            'target_margin': 0.40,
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual((self.data_dir / 'data.json').read_text(encoding='utf-8'), before)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response.headers.get('Content-Type', ''),
        )

    def test_dashboard_endpoint_returns_project_and_resource_analytics(self):
        response = self.client.get('/api/dashboard')

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload['kpis']['projects'], 1)
        self.assertEqual(payload['kpis']['resources'], 1)
        self.assertGreaterEqual(payload['kpis']['hours'], 8)
        self.assertEqual(payload['top_customers'][0]['label'], 'Saved Project')
        self.assertEqual(payload['top_roles_by_hours'][0]['label'], 'Architect')

    def test_all_projects_summary_includes_new_metadata_columns(self):
        response = self.client.get('/api/projects?summary=true')

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]['resource_count'], 1)
        self.assertEqual(payload[0]['total_hours'], 8.0)
        self.assertEqual(payload[0]['reference'], '')
        self.assertEqual(payload[0]['status'], '')
        self.assertEqual(payload[0]['currency'], 'USD')
        self.assertEqual(payload[0]['business_unit'], '')
        self.assertIn('costs', payload[0])

    def test_project_metadata_round_trips_through_save_and_load(self):
        new_project = {
            'project': {
                'customer': 'Round Trip Co',
                'location': 'India',
                'reference': 'REF-500',
                'proposal_date': '2026-04-08',
                'customer_first_touch_point': 'Referral',
                'project_description': 'Large transformation',
                'partner': 'AutomatonsX',
                'payment_terms': 'Net 30',
                'duration_months': 6,
                'status': 'Submitted',
                'stage': 'Proposal',
                'priority': 'High',
                'project_owner': 'Alice',
                'business_unit': 'AI',
                'account_manager': 'Bob',
                'sales_spoc': 'Carol',
                'delivery_manager': 'Dan',
                'technical_lead': 'Eve',
                'expected_start_date': '2026-05-01',
                'expected_end_date': '2026-10-31',
                'opportunity_id': 'OPP-500',
                'project_type': 'Implementation',
                'industry': 'Healthcare',
                'delivery_model': 'Hybrid',
                'billing_type': 'Fixed Bid',
                'currency': 'USD',
                'discount_pct': 5,
                'travel_cost': 1000,
                'infra_cost': 500,
                'third_party_cost': 250,
                'internal_notes': 'Internal note',
                'risks': 'Schedule risk',
                'dependencies': 'Client data',
                'next_action': 'Review commercials',
                'next_follow_up_date': '2026-04-15',
            },
            'resources': [{'role': 'Architect', 'level': 'L1', 'hours': 16}],
            'rate_card': [{'level': 'L1', 'rate': 25}],
            'role_catalog': [{'group': 'Delivery', 'roles': ['Architect']}],
            'business_units': ['EDM', 'AI', 'SAP'],
            'target_margin': 0.4,
            'fx_rate': 83.5,
        }

        create = self.client.post('/api/projects', json=new_project)
        self.assertEqual(create.status_code, 200)
        created = create.get_json()

        load = self.client.get(f"/api/projects/{created['id']}")
        self.assertEqual(load.status_code, 200)
        payload = load.get_json()

        self.assertEqual(payload['project']['reference'], 'REF-500')
        self.assertEqual(payload['project']['status'], 'Submitted')
        self.assertEqual(payload['project']['project_owner'], 'Alice')
        self.assertEqual(payload['project']['business_unit'], 'AI')
        self.assertEqual(payload['project']['currency'], 'USD')
        self.assertEqual(payload['project']['discount_pct'], 5)
        self.assertEqual(payload['project']['travel_cost'], 1000)
        self.assertEqual(payload['project']['next_follow_up_date'], '2026-04-15')
        self.assertEqual(payload['fx_rate'], 83.5)

        summary = self.client.get('/api/projects?summary=true').get_json()
        saved = next(item for item in summary if item['id'] == created['id'])
        self.assertEqual(saved['reference'], 'REF-500')
        self.assertEqual(saved['status'], 'Submitted')
        self.assertEqual(saved['project_owner'], 'Alice')
        self.assertEqual(saved['business_unit'], 'AI')
        self.assertEqual(saved['resource_count'], 1)
        self.assertEqual(saved['add_on_cost'], 1750.0)
        self.assertEqual(saved['fx_rate'], 83.5)

    def test_index_renders_new_project_fields(self):
        response = self.client.get('/')

        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('id="proj_status"', html)
        self.assertIn('id="proj_owner"', html)
        self.assertIn('id="proj_business_unit"', html)
        self.assertIn('id="proj_start_date"', html)
        self.assertIn('id="proj_discount_pct"', html)
        self.assertIn('id="proj_internal_notes"', html)
        self.assertIn('Application deployment refresh', html)

    def test_import_excel_supports_dynamic_project_and_resource_sheets(self):
        def build_workbook(wb):
            ws_meta = wb.active
            ws_meta.title = 'Intake'
            ws_meta.append(['Customer Name', 'Region', 'Project Owner', 'Business Unit', 'Currency'])
            ws_meta.append(['Dynamic Co', 'India', 'Alice', 'AI', 'USD'])
            ws_meta.append([])
            ws_meta.append(['Project Description', 'Billing Type', 'Expected Start Date'])
            ws_meta.append(['Migration Program', 'Fixed Bid', '2026-05-01'])

            ws_team = wb.create_sheet('Team Plan')
            ws_team.append(['Practice', 'Position', 'Grade', 'Planned Hours'])
            ws_team.append(['Delivery', 'Architect', 'L1', 12])
            ws_team.append(['QA', 'Tester', 'L2', 24])

        workbook = self._build_excel_bytes(build_workbook)

        with patch('pnl.routes.import_excel.load_global_settings', return_value={}), \
             patch('pnl.routes.import_excel.save_project_record') as save_project_record, \
             patch('pnl.routes.import_excel.save_project_version') as save_project_version:
            response = self.client.post(
                '/api/import-excel',
                data={'file': (workbook, 'dynamic-import.xlsx')},
                content_type='multipart/form-data',
            )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload['imported_count'], 1)
        self.assertEqual(len(payload['projects']), 1)
        self.assertEqual(payload['projects'][0]['project']['customer'], 'Dynamic Co')
        self.assertEqual(payload['projects'][0]['project']['project_owner'], 'Alice')
        self.assertEqual(payload['projects'][0]['project']['business_unit'], 'AI')
        self.assertEqual(payload['projects'][0]['project']['billing_type'], 'Fixed Bid')
        save_project_record.assert_called_once()
        save_project_version.assert_called_once()

    def test_import_excel_scans_multiple_sheets_when_not_native(self):
        def build_workbook(wb):
            ws_notes = wb.active
            ws_notes.title = 'Cover'
            ws_notes['A1'] = 'Upload generated from another tool'

            ws_project = wb.create_sheet('Project Summary')
            ws_project.append(['Client', 'Reference Number', 'Delivery Model'])
            ws_project.append(['Acme Labs', 'REF-777', 'Hybrid'])

            ws_resources = wb.create_sheet('Resource Dump')
            ws_resources.append(['Department', 'Job Role', 'Band', 'Effort Hours'])
            ws_resources.append(['SAP', 'Consultant', 'L3', 40])
            ws_resources.append(['SAP', 'Developer', 'L2', 32])

        workbook = self._build_excel_bytes(build_workbook)

        with patch('pnl.routes.import_excel.load_global_settings', return_value={}), \
             patch('pnl.routes.import_excel.save_project_record') as save_project_record, \
             patch('pnl.routes.import_excel.save_project_version') as save_project_version:
            response = self.client.post(
                '/api/import-excel',
                data={'file': (workbook, 'multi-sheet.xlsx')},
                content_type='multipart/form-data',
            )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload['imported_count'], 1)
        self.assertEqual(payload['projects'][0]['project']['customer'], 'Acme Labs')
        self.assertEqual(payload['projects'][0]['project']['reference'], 'REF-777')
        self.assertEqual(payload['projects'][0]['project']['delivery_model'], 'Hybrid')
        self.assertEqual(save_project_record.call_count, 1)
        self.assertEqual(save_project_version.call_count, 1)

    def test_import_excel_accepts_project_only_workbook(self):
        def build_workbook(wb):
            ws = wb.active
            ws.title = 'Project Intake'
            ws.append([
                'Company Name', 'Customer Name', 'Project Status', 'Stage', 'Priority',
                'Business Unit', 'Account Manager', 'Sales SPOC', 'Project Type',
                'Industry', 'Currency', 'Next Follow Up Date'
            ])
            ws.append([
                'AutomatonsX', 'Project Only Co', 'Submitted', 'Proposal', 'High',
                'SAP', 'Bob', 'Carol', 'Implementation',
                'Retail', 'USD', '2026-04-20'
            ])

        workbook = self._build_excel_bytes(build_workbook)

        with patch('pnl.routes.import_excel.load_global_settings', return_value={}), \
             patch('pnl.routes.import_excel.save_project_record') as save_project_record, \
             patch('pnl.routes.import_excel.save_project_version') as save_project_version:
            response = self.client.post(
                '/api/import-excel',
                data={'file': (workbook, 'project-only.xlsx')},
                content_type='multipart/form-data',
            )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload['imported_count'], 1)
        project = payload['projects'][0]['project']
        self.assertEqual(project['company'], 'AutomatonsX')
        self.assertEqual(project['customer'], 'Project Only Co')
        self.assertEqual(project['status'], 'Submitted')
        self.assertEqual(project['stage'], 'Proposal')
        self.assertEqual(project['priority'], 'High')
        self.assertEqual(project['business_unit'], 'SAP')
        self.assertEqual(project['account_manager'], 'Bob')
        self.assertEqual(project['sales_spoc'], 'Carol')
        self.assertEqual(project['project_type'], 'Implementation')
        self.assertEqual(project['industry'], 'Retail')
        self.assertEqual(project['currency'], 'USD')
        self.assertEqual(project['next_follow_up_date'], '2026-04-20')
        self.assertEqual(save_project_record.call_count, 1)
        self.assertEqual(save_project_version.call_count, 1)

    def test_import_excel_imports_multiple_project_rows_with_fuzzy_headers(self):
        def build_workbook(wb):
            ws = wb.active
            ws.title = 'Bulk Import'
            ws.append(['Client Name', 'Project Stat', 'Pipeline Stag', 'Proj Owner', 'Follow-up Date'])
            ws.append(['Alpha Corp', 'Draft', 'Discovery', 'Alice', '2026-04-22'])
            ws.append(['Beta Corp', 'Submitted', 'Proposal', 'Bob', '2026-04-29'])

        workbook = self._build_excel_bytes(build_workbook)

        with patch('pnl.routes.import_excel.load_global_settings', return_value={}), \
             patch('pnl.routes.import_excel.save_project_record') as save_project_record, \
             patch('pnl.routes.import_excel.save_project_version') as save_project_version:
            response = self.client.post(
                '/api/import-excel',
                data={'file': (workbook, 'bulk-projects.xlsx')},
                content_type='multipart/form-data',
            )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload['imported_count'], 2)
        self.assertEqual([item['project']['customer'] for item in payload['projects']], ['Alpha Corp', 'Beta Corp'])
        self.assertEqual(payload['projects'][0]['project']['status'], 'Draft')
        self.assertEqual(payload['projects'][0]['project']['stage'], 'Discovery')
        self.assertEqual(payload['projects'][0]['project']['project_owner'], 'Alice')
        self.assertEqual(payload['projects'][1]['project']['customer'], 'Beta Corp')
        self.assertEqual(save_project_record.call_count, 2)
        self.assertEqual(save_project_version.call_count, 2)


if __name__ == '__main__':
    unittest.main()
