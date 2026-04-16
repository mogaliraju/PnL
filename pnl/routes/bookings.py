"""Order Bookings & Commits CRUD routes."""
import re
from datetime import datetime
from io import BytesIO

import openpyxl
from flask import Blueprint, request, jsonify, session
from pnl.utils.storage import (
    list_order_bookings,
    save_order_booking,
    load_order_booking,
    delete_order_booking,
    load_custom_fields,
    save_custom_fields,
    load_column_labels,
    save_column_labels,
)
from pnl.utils.auth import login_required
from pnl.utils.logger import get_logger

bp = Blueprint('bookings', __name__)
log = get_logger(__name__)

BUS = ['SAP', 'EDM', 'AE', 'AI', 'MWP', 'RPA', 'Data']


@bp.route('/api/bookings', methods=['GET'])
@login_required
def list_bookings():
    booking_type = request.args.get('type')  # OTC or MRC
    return jsonify(list_order_bookings(booking_type))


@bp.route('/api/bookings', methods=['POST'])
@login_required
def create_booking():
    data = request.json or {}
    booking_id = 'booking_' + datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:24]
    data['saved_at'] = datetime.now().isoformat(timespec='seconds')
    data['saved_by'] = session.get('user', '')
    data['id'] = booking_id
    save_order_booking(booking_id, data)
    log.info(f"Booking '{booking_id}' ({data.get('booking_type','OTC')}) created by '{session.get('user')}'")
    return jsonify({'status': 'ok', 'id': booking_id})


@bp.route('/api/bookings/<booking_id>', methods=['GET'])
@login_required
def get_booking(booking_id):
    entry = load_order_booking(booking_id)
    if entry is None:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(entry)


@bp.route('/api/bookings/<booking_id>', methods=['PUT'])
@login_required
def update_booking(booking_id):
    existing = load_order_booking(booking_id)
    if existing is None:
        return jsonify({'error': 'Not found'}), 404
    data = request.json or {}
    data['id'] = booking_id
    data['saved_at'] = datetime.now().isoformat(timespec='seconds')
    data['saved_by'] = session.get('user', '')
    save_order_booking(booking_id, data)
    log.info(f"Booking '{booking_id}' updated by '{session.get('user')}'")
    return jsonify({'status': 'ok', 'id': booking_id})


@bp.route('/api/bookings/<booking_id>', methods=['DELETE'])
@login_required
def delete_booking(booking_id):
    delete_order_booking(booking_id)
    log.info(f"Booking '{booking_id}' deleted by '{session.get('user')}'")
    return jsonify({'status': 'ok'})


@bp.route('/api/bookings/meta/options')
@login_required
def bookings_options():
    return jsonify({'bus': BUS})


@bp.route('/api/bookings/meta/custom-fields', methods=['GET'])
@login_required
def get_bookings_custom_fields():
    return jsonify({
        'custom_fields': load_custom_fields('bookings'),
        'column_labels': load_column_labels('bookings'),
    })


@bp.route('/api/bookings/meta/custom-fields', methods=['POST'])
@login_required
def save_bookings_custom_fields():
    body = request.json or {}
    if 'custom_fields' in body:
        save_custom_fields('bookings', body['custom_fields'])
    if 'column_labels' in body:
        save_column_labels('bookings', body['column_labels'])
    return jsonify({'status': 'ok'})


# ── Bookings Import ───────────────────────────────────────────

# OTC sheet column aliases → internal field keys
_OTC_ALIASES = {
    'opf_number':           ('opf number', 'opf no', 'opf#', 'order number', 'po number'),
    'opf_date':             ('opf date', 'order date', 'booking date', 'opf raised date'),
    'cdd':                  ('cdd', 'contract delivery date', 'delivery date', 'commitment date'),
    'bu':                   ('bu', 'business unit', 'practice', 'vertical'),
    'customer_name':        ('customer name', 'customer', 'client', 'account name', 'account'),
    'otc':                  ('otc', 'one time cost', 'one-time cost', 'deal value', 'contract value'),
    'billed_pct':           ('billed %', 'billed pct', 'billing %', 'billed percentage', 'billed'),
    'milestones':           ('milestones', 'milestone', 'no of milestones', '# milestones'),
    'c4c_invoice_raised':   ('invoice raised from c4c', 'c4c invoice raised', 'c4c billed', 'c4c invoice'),
    'c4c_amount_received':  ('amount received from customer', 'c4c amount received', 'c4c collected', 'c4c received'),
    'c4c_pending_billing':  ('pending billing from c4c', 'c4c pending billing', 'c4c pending', 'pending billing'),
    'ax_invoice_raised':    ('invoice raised from ax', 'ax invoice raised', 'ax billed', 'ax invoice', 'automatonsx invoice'),
    'ax_amount_received':   ('amount received from c4c to ax', 'ax amount received', 'ax collected', 'ax received'),
    'billing_team_comments':('billing team comments', 'billing comments', 'billing team'),
    'pmo':                  ('pmo',),
}

# MRC sheet column aliases
_MRC_ALIASES = {
    'opf_number':           ('opf number', 'opf no', 'opf#', 'order number'),
    'opf_date':             ('opf date', 'order date', 'opf raised date'),
    'cdd':                  ('cdd', 'contract delivery date', 'delivery date'),
    'customer_name':        ('customer name', 'customer', 'client', 'account name', 'account'),
    'bu':                   ('bu', 'business unit', 'practice'),
    'mrc':                  ('mrc', 'monthly recurring cost', 'monthly recurring', 'monthly value'),
    'billed_pct':           ('billed %', 'billed pct', 'billing %', 'billed percentage'),
    'c4c_invoice_raised':   ('invoice raised from c4c', 'c4c invoice raised', 'c4c billed'),
    'c4c_amount_received':  ('amount received from customer', 'c4c amount received', 'c4c received'),
    'c4c_pending_billing':  ('pending billing from c4c', 'c4c pending billing', 'c4c pending'),
    'ax_invoice_raised':    ('invoice raised from ax', 'ax invoice raised', 'ax billed', 'ax invoice'),
    'ax_amount_received':   ('amount received from c4c to ax', 'ax amount received', 'ax received'),
    'ax_pending_collection':('pending collection from cloud4c', 'ax pending collection', 'pending collection'),
    'billing_team_comments':('billing team comments', 'billing comments', 'billing team'),
    'pmo':                  ('pmo',),
    'updates':              ('updates', 'notes', 'remarks'),
}


def _bk_norm(val):
    if val is None:
        return ''
    return ' '.join(re.sub(r'[^a-z0-9]+', ' ', str(val).lower().strip()).split())


def _bk_match_col(header, aliases_map):
    h = _bk_norm(header)
    if not h:
        return None
    # Pass 1: exact match
    for field, aliases in aliases_map.items():
        for alias in aliases:
            if h == _bk_norm(alias):
                return field
    # Pass 2: alias is a substring of header (e.g. 'otc' in 'sum of otc usd')
    for field, aliases in aliases_map.items():
        for alias in aliases:
            alias_n = _bk_norm(alias)
            if alias_n and alias_n in h:
                return field
    return None


def _bk_cell(val):
    import datetime as _dt
    if val is None:
        return None
    if isinstance(val, (_dt.datetime, _dt.date)):
        return val.strftime('%Y-%m-%d') if isinstance(val, _dt.datetime) else val.isoformat()
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    # Strip currency symbols and commas from numeric strings
    clean = re.sub(r'[$,\s]', '', s)
    try:
        return float(clean) if '.' in clean else int(clean)
    except ValueError:
        return s if s else None


def _detect_booking_header(ws, aliases_map, max_scan=20):
    """Find the header row with the most matching columns."""
    best = None
    best_count = 0
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True), start=1):
        mapping = {}
        for col_idx, cell in enumerate(row, start=1):
            field = _bk_match_col(cell, aliases_map)
            if field and field not in mapping:
                mapping[field] = col_idx
        if len(mapping) > best_count:
            best_count = len(mapping)
            best = (row_idx, mapping)
    return best if best_count >= 2 else None


def _parse_booking_sheet(ws, booking_type, aliases_map):
    result = _detect_booking_header(ws, aliases_map)
    if not result:
        return [], f'No recognisable {booking_type} columns.'
    header_row_idx, col_map = result
    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None or str(v).strip() in ('', '-', ' -   ') for v in row):
            continue
        entry = {}
        for field, col_idx in col_map.items():
            val = row[col_idx - 1] if col_idx - 1 < len(row) else None
            entry[field] = _bk_cell(val)
        # Need at least OPF number or customer to be a valid row
        if not entry.get('opf_number') and not entry.get('customer_name'):
            continue
        entry['booking_type'] = booking_type
        rows.append(entry)
    return rows, None


@bp.route('/api/bookings/import', methods=['POST'])
@login_required
def import_bookings():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'error': 'No file selected'}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return jsonify({'error': 'File must be an Excel workbook (.xlsx)'}), 400

    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), read_only=True, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Cannot open file: {e}'}), 400

    imported = []
    warnings = []

    # Sheet name hints for automatic type detection
    OTC_HINTS = ('otc', 'one time', 'one-time', 'booking revenue otc', 'otc closure')
    MRC_HINTS = ('mrc', 'monthly', 'recurring', 'mrc closure', 'booking revenue mrc')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        name_lower = sheet_name.lower()

        # Determine booking type from sheet name
        if any(h in name_lower for h in OTC_HINTS):
            btype, aliases = 'OTC', _OTC_ALIASES
        elif any(h in name_lower for h in MRC_HINTS):
            btype, aliases = 'MRC', _MRC_ALIASES
        else:
            # Try to auto-detect by seeing which alias set gives more matches
            otc_res = _detect_booking_header(ws, _OTC_ALIASES)
            mrc_res = _detect_booking_header(ws, _MRC_ALIASES)
            if not otc_res and not mrc_res:
                warnings.append(f'Sheet "{sheet_name}": no bookings columns found — skipped.')
                continue
            otc_count = len(otc_res[1]) if otc_res else 0
            mrc_count = len(mrc_res[1]) if mrc_res else 0
            if mrc_count > otc_count:
                btype, aliases = 'MRC', _MRC_ALIASES
            else:
                btype, aliases = 'OTC', _OTC_ALIASES

        rows, warn = _parse_booking_sheet(ws, btype, aliases)
        if warn:
            warnings.append(f'Sheet "{sheet_name}": {warn}')
        for entry in rows:
            booking_id = 'booking_' + datetime.now().strftime('%Y%m%d_%H%M%S_%f')
            entry['id'] = booking_id
            entry['saved_at'] = datetime.now().isoformat(timespec='seconds')
            entry['saved_by'] = session.get('user', '')
            save_order_booking(booking_id, entry)
            imported.append({
                'id': booking_id,
                'type': btype,
                'opf': entry.get('opf_number', ''),
                'customer': entry.get('customer_name', ''),
            })

    wb.close()

    if not imported:
        return jsonify({'error': 'No booking rows could be extracted.', 'warnings': warnings}), 400

    log.info(f"Bookings import: '{f.filename}' by '{session.get('user')}' — {len(imported)} rows")
    resp = {'imported_count': len(imported), 'rows': imported}
    if warnings:
        resp['warnings'] = warnings
    return jsonify(resp)
