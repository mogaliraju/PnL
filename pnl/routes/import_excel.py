"""Excel import route - bulk-imports project metadata from flexible Excel layouts."""
import re
from datetime import datetime
from difflib import SequenceMatcher
from io import BytesIO

import openpyxl
from flask import Blueprint, jsonify, request, session

from pnl.utils.auth import login_required
from pnl.utils.logger import get_logger
from pnl.utils.storage import load_global_settings, safe_filename, save_project_record, save_project_version

bp = Blueprint('import_excel', __name__)
log = get_logger(__name__)


_PROJECT_FIELD_ALIASES = {
    'company': ('company', 'company name', 'vendor', 'partner company', 'organization'),
    'customer': ('customer', 'customer name', 'client', 'client name', 'account', 'project name'),
    'location': ('location', 'country', 'region', 'geography', 'geo'),
    'reference': ('reference', 'reference no', 'reference no.', 'reference number', 'ref no', 'ref'),
    'proposal_date': ('proposal date', 'date', 'created date', 'submission date', 'proposal created date'),
    'customer_first_touch_point': ('customer first touch point', 'first touch point', 'first touch', 'lead source'),
    'duration_months': ('duration', 'duration months', 'duration in months', 'duration (months)', 'tenure'),
    'description': ('project description', 'description', 'scope', 'project scope', 'summary'),
    'partner': ('partner', 'partner details', 'partner name', 'implementation partner'),
    'payment_terms': ('payment terms', 'payment', 'terms', 'payment condition'),
    'status': ('status', 'project status', 'opportunity status'),
    'stage': ('stage', 'sales stage', 'pipeline stage', 'opportunity stage'),
    'priority': ('priority', 'project priority'),
    'project_owner': ('project owner', 'owner', 'project lead'),
    'business_unit': ('business unit', 'bu', 'practice'),
    'account_manager': ('account manager', 'account owner'),
    'sales_spoc': ('sales spoc', 'sales contact', 'sales owner', 'sales person'),
    'delivery_manager': ('delivery manager', 'delivery owner'),
    'technical_lead': ('technical lead', 'solution architect', 'architect', 'technical owner'),
    'expected_start_date': ('expected start date', 'start date', 'planned start date'),
    'expected_end_date': ('expected end date', 'end date', 'planned end date'),
    'opportunity_id': ('opportunity id', 'opportunity', 'crm id', 'opportunity number'),
    'project_type': ('project type', 'engagement type', 'type'),
    'industry': ('industry', 'vertical', 'sector'),
    'delivery_model': ('delivery model', 'delivery mode'),
    'billing_type': ('billing type', 'contract type', 'billing model'),
    'currency': ('currency',),
    'discount_pct': ('discount %', 'discount pct', 'discount percentage', 'discount'),
    'travel_cost': ('travel cost', 'travel'),
    'infra_cost': ('infra cost', 'infrastructure cost', 'infra'),
    'third_party_cost': ('third party cost', 'third-party cost', 'external cost'),
    'internal_notes': ('internal notes', 'notes', 'remarks', 'comment'),
    'risks': ('risks', 'risk'),
    'dependencies': ('dependencies', 'dependency'),
    'next_action': ('next action', 'action item', 'next step'),
    'next_follow_up_date': ('next follow up date', 'next follow-up date', 'follow up date', 'follow-up date'),
}
_HEADER_LABELS = {alias for aliases in _PROJECT_FIELD_ALIASES.values() for alias in aliases}
_HEADER_STOP_WORDS = {
    'total', 'subtotal', 'grand total', 'input cost', 'sell cost', 'revenue',
    'cost', 'markup', 'gross margin', 'summary',
}


def _cv(ws, row, col):
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


def _s(val):
    return str(val).strip() if val is not None else ''


def _n(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    try:
        f = float(str(val).strip().replace(',', ''))
        return int(f) if f == int(f) else f
    except Exception:
        return None


def _norm(val):
    return ' '.join(_s(val).lower().split()).rstrip(':')


def _norm_header(val):
    text = _norm(val).replace('\n', ' ')
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return ' '.join(text.split())


def _cell_to_value(raw):
    import datetime as _dt

    if raw is None:
        return None
    if isinstance(raw, _dt.datetime):
        return raw.strftime('%Y-%m-%d')
    if isinstance(raw, _dt.date):
        return raw.isoformat()
    if isinstance(raw, (int, float)):
        return _n(raw)
    text = _s(raw)
    return text or None


def _row_values(ws, row_idx, max_col=None):
    max_col = max_col or (ws.max_column or 1)
    return [_cv(ws, row_idx, col) for col in range(1, max_col + 1)]


def _looks_blank_row(values):
    return all(v in (None, '') for v in values)


def _token_similarity(a, b):
    at = set(_norm_header(a).split())
    bt = set(_norm_header(b).split())
    if not at or not bt:
        return 0.0
    return len(at & bt) / max(len(at), len(bt))


def _alias_score(header, alias):
    header_norm = _norm_header(header)
    alias_norm = _norm_header(alias)
    if not header_norm or not alias_norm:
        return 0.0
    if header_norm == alias_norm:
        return 1.0
    if alias_norm in header_norm or header_norm in alias_norm:
        return 0.96
    return max(
        SequenceMatcher(None, header_norm, alias_norm).ratio(),
        _token_similarity(header_norm, alias_norm),
    )


def _best_project_field(header):
    best_field = None
    best_score = 0.0
    for field, aliases in _PROJECT_FIELD_ALIASES.items():
        score = max(_alias_score(header, alias) for alias in aliases)
        if score > best_score:
            best_field = field
            best_score = score
    if best_score >= 0.74:
        return best_field, best_score
    return None, 0.0


def _find_label_value(ws, *label_variants, max_row=40, max_col=20):
    targets = {_norm_header(label) for label in label_variants}
    for r in range(1, min(ws.max_row or max_row, max_row) + 1):
        for c in range(1, min(ws.max_column or max_col, max_col) + 1):
            if _norm_header(_cv(ws, r, c)) not in targets:
                continue
            for rr, cc in ((r, c + 1), (r + 1, c)):
                value = _cv(ws, rr, cc)
                if value is not None and _s(value) and _norm_header(value) not in _HEADER_LABELS:
                    return value
    return None


def _parse_native_project_sheet(ws):
    return {
        'customer': _cell_to_value(_find_label_value(ws, 'Customer Name', 'Customer')),
        'location': _cell_to_value(_find_label_value(ws, 'Location')),
        'reference': _cell_to_value(_find_label_value(ws, 'Reference No.', 'Reference No', 'Reference Number', 'Reference')),
        'proposal_date': _cell_to_value(_find_label_value(ws, 'Proposal Date', 'Date')),
        'duration_months': _n(_find_label_value(ws, 'Duration (Months)', 'Duration')),
        'description': _cell_to_value(_find_label_value(ws, 'Project Description')),
        'partner': _cell_to_value(_find_label_value(ws, 'Partner Details', 'Partner')),
        'payment_terms': _cell_to_value(_find_label_value(ws, 'Payment Terms')),
    }


def _detect_project_header_rows(ws, max_scan_rows=50):
    max_col = min(ws.max_column or 25, 80)
    candidates = []
    for row_idx in range(1, min(ws.max_row or max_scan_rows, max_scan_rows) + 1):
        field_map = {}
        score = 0.0
        for col_idx, raw in enumerate(_row_values(ws, row_idx, max_col), start=1):
            if _s(raw) == '':
                continue
            field, field_score = _best_project_field(raw)
            if not field:
                continue
            current = field_map.get(field)
            if current and current['score'] >= field_score:
                continue
            field_map[field] = {'col': col_idx, 'score': field_score}
        if 'customer' not in field_map or len(field_map) < 2:
            continue
        avg_score = sum(item['score'] for item in field_map.values()) / len(field_map)
        customer_score = field_map['customer']['score']
        if customer_score < 0.9 or avg_score < 0.84:
            continue
        score = sum(item['score'] for item in field_map.values())
        candidates.append({
            'row': row_idx,
            'fields': {field: item['col'] for field, item in field_map.items()},
            'score': score,
        })

    deduped = []
    seen = set()
    for candidate in sorted(candidates, key=lambda item: (-len(item['fields']), -item['score'], item['row'])):
        sig = tuple(sorted(candidate['fields'].items()))
        if sig in seen:
            continue
        seen.add(sig)
        deduped.append(candidate)
    return deduped


def _project_value_from_row(fields, row):
    project = {}
    for field, idx in fields.items():
        if idx - 1 >= len(row):
            continue
        value = _cell_to_value(row[idx - 1])
        if value is None:
            continue
        project[field] = value
    return project


def _merge_project_data(base, update):
    for key, value in update.items():
        if value not in (None, '') and not base.get(key):
            base[key] = value
    return base


def _row_looks_like_project_header(row):
    nonblank = [_s(value) for value in row if _s(value)]
    if not nonblank:
        return False
    header_hits = 0
    for value in nonblank:
        field, score = _best_project_field(value)
        if field and score >= 0.9:
            header_hits += 1
    return header_hits >= 2 and (header_hits / len(nonblank)) >= 0.6


def _looks_like_summary_row(project):
    customer = _norm_header(project.get('customer'))
    if not customer:
        return False
    return customer in _HEADER_STOP_WORDS or customer in {'customer', 'client', 'project name'}


def _normalize_project(project):
    normalized = {k: v for k, v in project.items() if v not in (None, '')}
    if 'description' in normalized:
        normalized['project_description'] = normalized.pop('description')
    for numeric_field in ('duration_months', 'discount_pct', 'travel_cost', 'infra_cost', 'third_party_cost'):
        if numeric_field in normalized:
            normalized[numeric_field] = _n(normalized[numeric_field])
    for text_field, default in {
        'company': 'AutomatonsX',
        'partner': 'AutomatonsX',
        'payment_terms': 'As per proposal',
        'status': 'Draft',
        'stage': 'Qualification',
        'priority': 'Medium',
        'delivery_model': 'Offshore',
        'billing_type': 'Time & Material',
        'currency': 'USD',
    }.items():
        normalized[text_field] = _s(normalized.get(text_field) or default)
    normalized['customer'] = _s(normalized.get('customer'))
    return normalized


def _parse_project_rows(ws, header_info):
    projects = []
    fields = header_info['fields']
    max_col = min(ws.max_column or 25, 80)
    blank_streak = 0
    for row_idx in range(header_info['row'] + 1, min((ws.max_row or 1000), 5000) + 1):
        row = _row_values(ws, row_idx, max_col)
        if _looks_blank_row(row):
            blank_streak += 1
            if blank_streak >= 5:
                break
            continue
        blank_streak = 0
        if _row_looks_like_project_header(row):
            break
        project = _project_value_from_row(fields, row)
        if not project:
            continue
        if _looks_like_summary_row(project):
            break
        if not project.get('customer'):
            continue
        projects.append(project)
    return projects


def _parse_sheet_metadata(ws):
    max_col = min(ws.max_column or 25, 80)
    max_row = min(ws.max_row or 50, 50)
    merged = {}
    for row_idx in range(1, max_row):
        headers = {}
        for col_idx, raw in enumerate(_row_values(ws, row_idx, max_col), start=1):
            field, score = _best_project_field(raw)
            if field and score >= 0.9 and field not in headers:
                headers[field] = col_idx
        if len(headers) < 2:
            continue
        next_row = _row_values(ws, row_idx + 1, max_col)
        if _looks_blank_row(next_row) or _row_looks_like_project_header(next_row):
            continue
        _merge_project_data(merged, _project_value_from_row(headers, next_row))
    return merged


def _dedupe_projects(projects):
    deduped = []
    seen = set()
    for project in projects:
        key = (
            _norm_header(project.get('customer')),
            _norm_header(project.get('reference')),
            _norm_header(project.get('opportunity_id')),
            _norm_header(project.get('proposal_date')),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(project)
    return deduped


def _build_project_payload(project, settings, pid, name):
    return {
        '_meta': {
            'name': name,
            'id': pid,
            'saved_at': datetime.now().isoformat(timespec='seconds'),
            'saved_by': session.get('user', ''),
        },
        'project': project,
        'resources': [],
        'pnl_roles': [],
        'releases': [],
        'rate_card': settings.get('rate_card', []),
        'role_catalog': settings.get('role_catalog', []),
        'business_units': settings.get('business_units', []),
        'attachments': {'customer_po': False, 'cloud4c_quote': False, 'partner_proposal': False},
        'funding': {
            'marketing': {'currency': 'USD', 'value': None},
            'management': {'currency': 'USD', 'value': None},
            'discount': {'currency': 'USD', 'value': None},
        },
        'approvals': {'prepared_by': '', 'reviewed_by': '', 'approved_by': ''},
        'export_filename': '',
        'target_margin': 0.40,
        'fx_rate': None,
    }


def _persist_projects(projects):
    settings = load_global_settings()
    imported = []
    for index, project in enumerate(projects, start=1):
        name = project.get('customer') or project.get('reference') or f'Imported Project {index}'
        pid = safe_filename(name) + '_' + datetime.now().strftime('%Y%m%d_%H%M%S_%f')
        payload = _build_project_payload(project, settings, pid, name)
        save_project_record(pid, payload)
        save_project_version(pid, datetime.now().strftime('%Y%m%d_%H%M%S_%f'), payload)
        imported.append({'id': pid, 'name': name, 'project': project})
    return imported


@bp.route('/api/import-excel', methods=['POST'])
@login_required
def import_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'error': 'No file selected'}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return jsonify({'error': 'File must be an Excel workbook (.xlsx)'}), 400

    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), read_only=True, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Cannot open file: {e}'}), 400

    warnings = []
    projects = []
    sheet_names = wb.sheetnames
    sheet_names_lower = [s.lower() for s in sheet_names]
    pnl_sheet = next((sheet_names[i] for i, s in enumerate(sheet_names_lower) if s == 'pnl'), None)

    if pnl_sheet:
        try:
            native_project = _normalize_project(_parse_native_project_sheet(wb[pnl_sheet]))
            if native_project.get('customer'):
                projects.append(native_project)
        except Exception as e:
            warnings.append(f'PnL sheet error: {e}')

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        try:
            header_candidates = _detect_project_header_rows(ws)
            sheet_projects = []
            for candidate in header_candidates:
                sheet_projects.extend(_parse_project_rows(ws, candidate))
            sheet_metadata = _parse_sheet_metadata(ws)
            if len(sheet_projects) == 1 and sheet_metadata:
                _merge_project_data(sheet_projects[0], sheet_metadata)
            elif not sheet_projects and sheet_metadata.get('customer'):
                sheet_projects.append(sheet_metadata)
            projects.extend(
                _normalize_project(project)
                for project in sheet_projects
                if project.get('customer')
            )
        except Exception as e:
            warnings.append(f'Project scan failed on "{sheet_name}": {e}')

    wb.close()

    projects = _dedupe_projects(projects)
    if not projects:
        return jsonify({
            'error': 'Could not extract any project rows from the file.',
            'warnings': warnings,
        }), 400

    imported = _persist_projects(projects)
    log.info(
        f"import-excel: '{f.filename}' by '{session.get('user')}' imported_projects={len(imported)}"
    )

    response = {
        'imported_count': len(imported),
        'projects': imported,
    }
    if warnings:
        response['warnings'] = warnings
    return jsonify(response), 200
