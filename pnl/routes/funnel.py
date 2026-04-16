"""Funnel Report CRUD routes."""
import re
from datetime import datetime
from io import BytesIO

import openpyxl
from flask import Blueprint, request, jsonify, session
from pnl.utils.storage import (
    list_funnel_entries,
    save_funnel_entry,
    load_funnel_entry,
    delete_funnel_entry,
    safe_filename,
    load_custom_fields,
    save_custom_fields,
    load_column_labels,
    save_column_labels,
)
from pnl.utils.auth import login_required
from pnl.utils.logger import get_logger

bp = Blueprint('funnel', __name__)
log = get_logger(__name__)

STAGES = ['Identified', 'Qualified', 'Proposal Sent', 'Negotiation', 'Closed Won', 'Closed Lost', 'Pipeline']
PRODUCTS = ['Data', 'SAP Services', 'Echo', 'AE', 'AI', 'Connect / Echo', 'MWP', 'RPA']
REGIONS = ['India', 'MEST', 'ASEAN', 'ROW', 'APAC', 'North America', 'South America', 'Europe']
NET_FORECASTING = ['Pipeline', 'Upside', 'Best Case', 'Commit']


@bp.route('/api/funnel', methods=['GET'])
@login_required
def list_funnel():
    return jsonify(list_funnel_entries())


@bp.route('/api/funnel', methods=['POST'])
@login_required
def create_funnel():
    data = request.json or {}
    entry_id = 'funnel_' + datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:22]
    data['saved_at'] = datetime.now().isoformat(timespec='seconds')
    data['saved_by'] = session.get('user', '')
    data['id'] = entry_id
    save_funnel_entry(entry_id, data)
    log.info(f"Funnel entry '{entry_id}' created by '{session.get('user')}'")
    return jsonify({'status': 'ok', 'id': entry_id})


@bp.route('/api/funnel/<entry_id>', methods=['GET'])
@login_required
def get_funnel(entry_id):
    entry = load_funnel_entry(entry_id)
    if entry is None:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(entry)


@bp.route('/api/funnel/<entry_id>', methods=['PUT'])
@login_required
def update_funnel(entry_id):
    existing = load_funnel_entry(entry_id)
    if existing is None:
        return jsonify({'error': 'Not found'}), 404
    data = request.json or {}
    data['id'] = entry_id
    data['saved_at'] = datetime.now().isoformat(timespec='seconds')
    data['saved_by'] = session.get('user', '')
    save_funnel_entry(entry_id, data)
    log.info(f"Funnel entry '{entry_id}' updated by '{session.get('user')}'")
    return jsonify({'status': 'ok', 'id': entry_id})


@bp.route('/api/funnel/<entry_id>', methods=['DELETE'])
@login_required
def delete_funnel(entry_id):
    delete_funnel_entry(entry_id)
    log.info(f"Funnel entry '{entry_id}' deleted by '{session.get('user')}'")
    return jsonify({'status': 'ok'})


@bp.route('/api/funnel/meta/options')
@login_required
def funnel_options():
    return jsonify({
        'stages': STAGES,
        'products': PRODUCTS,
        'regions': REGIONS,
        'net_forecasting': NET_FORECASTING,
    })


@bp.route('/api/funnel/meta/custom-fields', methods=['GET'])
@login_required
def get_funnel_custom_fields():
    return jsonify({
        'custom_fields': load_custom_fields('funnel'),
        'column_labels': load_column_labels('funnel'),
    })


@bp.route('/api/funnel/meta/custom-fields', methods=['POST'])
@login_required
def save_funnel_custom_fields():
    body = request.json or {}
    if 'custom_fields' in body:
        save_custom_fields('funnel', body['custom_fields'])
    if 'column_labels' in body:
        save_column_labels('funnel', body['column_labels'])
    return jsonify({'status': 'ok'})


# ── Funnel Import ─────────────────────────────────────────────

# Maps Excel column header variations → internal field key
_FUNNEL_ALIASES = {
    'record_id':         ('record id', 'record_id', 'crm id', 'id'),
    'reporting_manager': ('reporting manager', 'manager', 'rm'),
    'opportunity_owner': ('opportunity owner', 'owner', 'sales owner', 'ae'),
    'region':            ('region', 'geo', 'geography', 'territory'),
    'account_name':      ('account name', 'account', 'company', 'customer', 'client'),
    'description':       ('what are they into', 'description', 'about', 'industry detail', 'business description'),
    'opportunity_name':  ('opportunity name', 'opportunity', 'project name', 'deal name'),
    'closing_month':     ('closing month', 'close month', 'target close', 'expected close month'),
    'ageing_days':       ('ageing days', 'ageing', 'aging', 'age days', 'days'),
    'stage':             ('stage', 'sales stage', 'pipeline stage', 'opportunity stage'),
    'fq':                ('fq', 'fiscal quarter', 'quarter', 'fiscal q'),
    'final_product':     ('final product', 'product', 'solution', 'product family', 'offering'),
    'created_time':      ('created time', 'created date', 'created at', 'creation date'),
    'net_forecasting':   ('net forecasting', 'forecasting', 'forecast category', 'net forecast'),
    'otc_usd_k':         ('sum of otc', 'otc', 'sum of otc (in usd k)', 'otc usd k', 'otc (usd k)'),
    'tcv_usd':           ('sum of tcv', 'tcv', 'sum of tcv ( usd)', 'tcv usd', 'total contract value'),
    'mrc_usd_k':         ('sum of mrc', 'mrc', 'sum of mrc (in usd k)', 'mrc usd k', 'mrc (usd k)'),
    'acv_usd_k':         ('sum of acv', 'acv', 'sum of acv (in usd k)', 'acv usd k', 'acv (usd k)', 'annual contract value'),
    'updates':           ('updates', 'notes', 'comments', 'update', 'remarks'),
}


def _fn_norm(val):
    if val is None:
        return ''
    return ' '.join(re.sub(r'[^a-z0-9]+', ' ', str(val).lower().strip()).split())


def _fn_match_col(header):
    h = _fn_norm(header)
    if not h:
        return None
    # Pass 1: exact match
    for field, aliases in _FUNNEL_ALIASES.items():
        for alias in aliases:
            if h == _fn_norm(alias):
                return field
    # Pass 2: alias is a substring of header
    for field, aliases in _FUNNEL_ALIASES.items():
        for alias in aliases:
            alias_n = _fn_norm(alias)
            if alias_n and alias_n in h:
                return field
    return None


def _fn_cell(val):
    import datetime as _dt
    if val is None:
        return None
    if isinstance(val, (_dt.datetime, _dt.date)):
        return val.strftime('%Y-%m-%d') if isinstance(val, _dt.datetime) else val.isoformat()
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    return s if s else None


@bp.route('/api/funnel/import', methods=['POST'])
@login_required
def import_funnel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'error': 'No file selected'}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return jsonify({'error': 'File must be an Excel workbook (.xlsx)'}), 400

    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), read_only=True, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Cannot open file: {e}'}), 400

    imported = []
    warnings = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find the header row (first row where we can map ≥3 funnel fields)
        header_row_idx = None
        col_map = {}
        for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
            mapping = {}
            for col_idx, cell in enumerate(row, start=1):
                field = _fn_match_col(cell)
                if field and field not in mapping:
                    mapping[field] = col_idx
            if len(mapping) >= 3:
                header_row_idx = row_idx
                col_map = mapping
                break

        if not col_map:
            warnings.append(f'Sheet "{sheet_name}": no recognisable funnel columns — skipped.')
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(v is None or str(v).strip() == '' for v in row):
                continue
            entry = {}
            for field, col_idx in col_map.items():
                val = row[col_idx - 1] if col_idx - 1 < len(row) else None
                entry[field] = _fn_cell(val)

            # Skip rows with no identifying information
            if not entry.get('account_name') and not entry.get('opportunity_name') and not entry.get('opportunity_owner'):
                continue

            entry_id = 'funnel_' + datetime.now().strftime('%Y%m%d_%H%M%S_%f')
            entry['id'] = entry_id
            entry['saved_at'] = datetime.now().isoformat(timespec='seconds')
            entry['saved_by'] = session.get('user', '')
            save_funnel_entry(entry_id, entry)
            imported.append({'id': entry_id, 'account': entry.get('account_name', ''), 'opportunity': entry.get('opportunity_name', '')})

    wb.close()

    if not imported:
        return jsonify({'error': 'No funnel rows could be extracted.', 'warnings': warnings}), 400

    log.info(f"Funnel import: '{f.filename}' by '{session.get('user')}' — {len(imported)} rows")
    resp = {'imported_count': len(imported), 'rows': imported}
    if warnings:
        resp['warnings'] = warnings
    return jsonify(resp)
