"""Excel workbook builder — isolated from Flask."""
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)


def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _set(ws, row, col, value, bold=False, bg=None, color='000000', num_fmt=None):
    c = ws.cell(row, col)
    c.value = value
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(vertical='center', wrap_text=True)
    if bg:
        c.fill = fill(bg)
    if num_fmt:
        c.number_format = num_fmt
    c.border = thin_border()
    return c


def _setc(ws, row, col, value, bold=False, bg=None, num_fmt=None):
    c = ws.cell(row, col)
    c.value = value
    c.font = Font(bold=bold)
    c.alignment = Alignment(vertical='center')
    if bg:
        c.fill = fill(bg)
    if num_fmt:
        c.number_format = num_fmt
    c.border = thin_border()
    return c


def build_workbook(data: dict, costs: dict) -> openpyxl.Workbook:
    input_cost = costs['input_cost']
    sell_cost  = costs['sell_cost']
    markup     = costs['markup']
    markup_pct = costs['markup_pct']
    gross_margin = costs['gross_margin']

    rate_map   = {r['level']: r['rate'] for r in data.get('rate_card', [])}
    resources  = data.get('resources', [])
    proj       = data.get('project', {})
    att        = data.get('attachments', {})
    fund       = data.get('funding', {})
    appr       = data.get('approvals', {})

    wb = openpyxl.Workbook()

    # ── Sheet 1: PnL ────────────────────────────────────────────
    ws = wb.active
    ws.title = 'PnL'

    for i, w in enumerate([22,18,22,20,18,12,14,12,12,14,14,10,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header rows
    ws.merge_cells('A1:M1')
    c = ws['A1']
    c.value = proj.get('company', 'AutomatonsX')
    c.font = Font(bold=True, size=16, color='FFFFFF')
    c.fill = fill('2D1B69')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:M2')
    c = ws['A2']
    c.value = 'Project Profit and Loss Statement'
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.fill = fill('6D28D9')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # Info rows 3-6
    info = [
        (3, 'Customer Name', proj.get('customer',''), 'Date', ''),
        (4, 'Location',      proj.get('location',''), 'Reference', proj.get('reference','')),
    ]
    for row, lbl, val, lbl2, val2 in info:
        _set(ws, row, 1, lbl, bold=True, bg='EDE9FE')
        ws.merge_cells(f'B{row}:J{row}')
        _set(ws, row, 2, val)
        _set(ws, row, 11, lbl2, bold=True, bg='EDE9FE')
        ws.merge_cells(f'L{row}:M{row}')
        _set(ws, row, 12, val2)

    _set(ws, 5, 1, 'Proposal Date', bold=True, bg='EDE9FE')
    ws.merge_cells('B5:J5')
    _set(ws, 5, 2, proj.get('proposal_date') or datetime.today().strftime('%Y-%m-%d'))
    ws.merge_cells('K5:M5')
    _set(ws, 5, 11, proj.get('customer_first_touch_point', ''))

    _set(ws, 6, 1, 'Proposal Value', bold=True, bg='EDE9FE')
    ws.merge_cells('B6:J6')
    _set(ws, 6, 2, sell_cost, num_fmt='#,##0.00')

    # Table header row 7
    ws.row_dimensions[7].height = 30
    hdrs = [('A7','Project Description'),('C7','Delivery Method'),('D7','Partner Details'),
            ('E7','Payment Terms'),('F7','Duration\n(Months)'),('G7','Revenue (USD)'),
            ('H7','Revenue Split'),('I7','Cost Split (%)'),('J7','Cost (USD)'),
            ('K7','Markup Value\n(USD)'),('L7','Markup %'),('M7','Gross Margin')]
    for ref, val in hdrs:
        c = ws[ref]
        c.value = val
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = fill('6D28D9')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.merge_cells('A7:B7')

    # PnL role rows
    pnl_roles = data.get('pnl_roles', [])
    for idx, role in enumerate(pnl_roles):
        r = 8 + idx
        ws.merge_cells(f'A{r}:B{r}')
        _set(ws, r, 1,
             proj.get('project_description', proj.get('customer','')) if idx == 0 else role.get('name',''),
             bold=(idx==0), bg='F5F3FF' if idx==0 else 'F9F9F9')
        _set(ws, r, 3, role.get('name',''))
        _set(ws, r, 4, role.get('partner',''))
        _set(ws, r, 5, role.get('payment_terms',''))
        _set(ws, r, 6, proj.get('duration_months',''))
        if idx == 0:
            _set(ws, r, 7,  sell_cost,   num_fmt='#,##0.00', bg='F0FDF4')
            _set(ws, r, 10, input_cost,  num_fmt='#,##0.00', bg='FEF3C7')
            _set(ws, r, 11, markup,      num_fmt='#,##0.00', bg='F0FDF4')
            _set(ws, r, 12, markup_pct,  num_fmt='0.0%')
            _set(ws, r, 13, gross_margin,num_fmt='0.0%')
        for col in range(1, 14):
            ws.cell(r, col).border = thin_border()

    end_row = 8 + len(pnl_roles) - 1

    # Attachments
    ar = end_row + 2
    _set(ws, ar, 3, 'Yes/No',                  bold=True, bg='EDE9FE')
    _set(ws, ar, 4, 'Reason, if not attached',  bold=True, bg='EDE9FE')
    for i, (lbl, val) in enumerate([
        ('Customer PO attached',    att.get('customer_po', False)),
        ('Cloud4C Quote attached',  att.get('cloud4c_quote', False)),
        ('Partner proposal attached', att.get('partner_proposal', False))
    ]):
        _set(ws, ar+1+i, 1, lbl, bold=True)
        _set(ws, ar+1+i, 3, 'Yes' if val else 'No')

    # Funding
    fr = ar + 5
    _set(ws, fr, 1, 'Funding / Rebates', bold=True, bg='6D28D9', color='FFFFFF')
    for col_l, h in [('C','Method of recovery'),('D','Reference'),('E','Currency'),('F','Value')]:
        c = ws[f'{col_l}{fr}']
        c.value = h; c.font = Font(bold=True, color='FFFFFF')
        c.fill = fill('6D28D9'); c.border = thin_border()
    fund_total = 0
    for i, (lbl, fk) in enumerate([('Marketing','marketing'),('Management','management'),('Discount','discount')]):
        fd = fund.get(fk, {})
        v  = fd.get('value') or 0
        fund_total += v
        _set(ws, fr+1+i, 1, lbl, bold=True)
        _set(ws, fr+1+i, 3, fd.get('method',''))
        _set(ws, fr+1+i, 4, fd.get('reference',''))
        _set(ws, fr+1+i, 5, fd.get('currency','USD'))
        _set(ws, fr+1+i, 6, v or '', num_fmt='#,##0.00' if v else None)
    _set(ws, fr+4, 5, 'Total', bold=True)
    _set(ws, fr+4, 6, fund_total, bold=True, bg='FEF3C7', num_fmt='#,##0.00')

    # Approvals
    xr = fr + 6
    for i, (lbl, val) in enumerate([
        ('Prepared By', appr.get('prepared_by','')),
        ('Reviewed By', appr.get('reviewed_by','')),
        ('Approved By', appr.get('approved_by',''))
    ]):
        _set(ws, xr+i, 1, lbl, bold=True, bg='EDE9FE')
        ws.merge_cells(f'B{xr+i}:G{xr+i}')
        _set(ws, xr+i, 2, val)

    # ── Sheet 2: Input Costing Calculation ─────────────────────
    wc = wb.create_sheet('Input Costing Calculation')
    for i, w in enumerate([12,12,14,8,22,6,22,6,10,6,10,10], 1):
        wc.column_dimensions[get_column_letter(i)].width = w

    rel_hdrs = ['Release','Entities','','EC+Time Off','Payroll Countries',
                'EC Payroll','ABAP, CPI, PTP','BTP','RCM+ONB+OFFB','PMGM','LMS','Months']
    for col, h in enumerate(rel_hdrs, 1):
        c = wc.cell(1, col); c.value = h
        if h:
            c.font = Font(bold=True, color='FFFFFF', size=9)
            c.fill = fill('2D1B69')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()

    releases = data.get('releases', [])
    totals = {k: 0 for k in ['entities','ec_time_off','ec_payroll','abap_cpi_ptp','btp','rcm_onb_offb','pmgm','lms']}
    countries = []
    for ri, rel in enumerate(releases):
        row = 2 + ri
        vals = [(1,rel['name']),(2,rel.get('entities')),(4,rel.get('ec_time_off')),
                (5,rel.get('payroll_countries','')),(6,rel.get('ec_payroll')),
                (7,rel.get('abap_cpi_ptp')),(8,rel.get('btp')),
                (9,rel.get('rcm_onb_offb')),(10,rel.get('pmgm')),(11,rel.get('lms'))]
        for col, v in vals:
            wc.cell(row, col).value = v
            wc.cell(row, col).border = thin_border()
        for k in totals:
            totals[k] += rel.get(k) or 0
        if rel.get('payroll_countries'):
            countries.append(rel['payroll_countries'])

    tr = 2 + len(releases)
    wc.cell(tr,1).value='Total'; wc.cell(tr,2).value=totals['entities']
    wc.cell(tr,4).value=totals['ec_time_off']; wc.cell(tr,5).value=', '.join(countries)
    wc.cell(tr,6).value=totals['ec_payroll']; wc.cell(tr,7).value=totals['abap_cpi_ptp']
    wc.cell(tr,8).value=totals['btp']; wc.cell(tr,9).value=totals['rcm_onb_offb']
    wc.cell(tr,10).value=totals['pmgm']; wc.cell(tr,11).value=totals['lms']
    wc.cell(tr,12).value=sum(totals[k] for k in ['ec_time_off','ec_payroll','abap_cpi_ptp','btp','rcm_onb_offb','pmgm','lms'])
    for col in range(1,13):
        wc.cell(tr,col).font=Font(bold=True); wc.cell(tr,col).fill=fill('EDE9FE'); wc.cell(tr,col).border=thin_border()

    sr = tr + 2
    rate_card = data.get('rate_card', [])
    _setc(wc,sr,1,'HOURS',bold=True,bg='FEF9C3'); _setc(wc,sr,3,'Cost',bold=True,bg='DCFCE7')
    _setc(wc,sr,4,'Level',bold=True,bg='EDE9FE'); _setc(wc,sr,5,'Role',bold=True,bg='EDE9FE')
    _setc(wc,sr,7,'Role',bold=True,bg='F3E8FF');  _setc(wc,sr,8,'Level',bold=True,bg='F3E8FF')
    _setc(wc,sr,9,'Hours',bold=True,bg='F3E8FF'); _setc(wc,sr,11,'Level',bold=True,bg='F0FDF4')
    _setc(wc,sr,12,'Rate ($/hr)',bold=True,bg='F0FDF4')

    rr = sr + 1
    total_h = total_c = 0
    for ri, res in enumerate(resources):
        rate = rate_map.get(res.get('level',''), 0)
        cost = (res.get('hours') or 0) * rate
        total_h += res.get('hours') or 0; total_c += cost
        _setc(wc,rr+ri,1,res.get('hours')); _setc(wc,rr+ri,3,round(cost,2),num_fmt='#,##0.00')
        _setc(wc,rr+ri,4,res.get('level')); _setc(wc,rr+ri,5,res.get('role'))
        _setc(wc,rr+ri,7,res.get('role'));  _setc(wc,rr+ri,8,res.get('level'))
        _setc(wc,rr+ri,9,res.get('hours'))
        if ri < len(rate_card):
            _setc(wc,rr+ri,11,rate_card[ri]['level'],bg='F0FDF4')
            _setc(wc,rr+ri,12,rate_card[ri]['rate'],bg='F0FDF4',num_fmt='#,##0.00')

    for ri in range(len(resources), len(rate_card)):
        _setc(wc,rr+ri,11,rate_card[ri]['level'],bg='F0FDF4')
        _setc(wc,rr+ri,12,rate_card[ri]['rate'],bg='F0FDF4',num_fmt='#,##0.00')

    mr = rr + max(len(resources), len(rate_card))
    _setc(wc,mr,1,total_h,bold=True,bg='FEF9C3')
    _setc(wc,mr,3,round(total_c,2),bold=True,bg='DCFCE7',num_fmt='#,##0.00')
    _setc(wc,mr,8,'Total',bold=True); _setc(wc,mr,9,total_h,bold=True,bg='FEF9C3')

    xr2 = mr + 2
    wc.cell(xr2,2).value = 'proposal'
    _setc(wc,xr2,3,'Input Cost',bold=True,bg='FEF3C7')
    _setc(wc,xr2,4,input_cost,num_fmt='#,##0.00',bg='FEF3C7')
    _setc(wc,xr2+1,3,'Sell Cost',bold=True,bg='F0FDF4')
    _setc(wc,xr2+1,4,sell_cost,num_fmt='#,##0.00',bg='F0FDF4')

    return wb
