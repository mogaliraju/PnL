import json
import os
import re
from datetime import datetime
from io import BytesIO
from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

app = Flask(__name__)

BASE_DIR      = os.path.dirname(__file__)
# Use /data (Render persistent disk) when available, else local directory
DATA_DIR      = os.environ.get('DATA_DIR', BASE_DIR)
DATA_FILE     = os.path.join(DATA_DIR, 'data.json')
SETTINGS_FILE = os.path.join(DATA_DIR, 'settings.json')
PROJECTS_DIR  = os.path.join(DATA_DIR, 'projects')
os.makedirs(PROJECTS_DIR, exist_ok=True)

# Seed data.json from bundled copy if not yet present on disk
_SEED_FILE = os.path.join(BASE_DIR, 'data.json')
if not os.path.exists(DATA_FILE) and os.path.exists(_SEED_FILE) and DATA_DIR != BASE_DIR:
    import shutil
    shutil.copy(_SEED_FILE, DATA_FILE)


def load_settings():
    if not os.path.exists(SETTINGS_FILE):
        return {}
    with open(SETTINGS_FILE, 'r') as f:
        return json.load(f)


def save_settings(s):
    with open(SETTINGS_FILE, 'w') as f:
        json.dump(s, f, indent=2)


def load_data():
    with open(DATA_FILE, 'r') as f:
        data = json.load(f)
    # Always inject the global standard rate card + role catalog
    s = load_settings()
    if s.get('rate_card'):
        data['rate_card'] = s['rate_card']
    if s.get('role_catalog'):
        data['role_catalog'] = s['role_catalog']
    return data


def save_data(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2)
    # Persist rate card & role catalog as global standards
    s = load_settings()
    if data.get('rate_card'):
        s['rate_card'] = data['rate_card']
    if data.get('role_catalog'):
        s['role_catalog'] = data['role_catalog']
    save_settings(s)


def safe_filename(name):
    return re.sub(r'[^a-zA-Z0-9_\-]', '_', name.strip())


# ── Current working project ──────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/data', methods=['GET'])
def get_data():
    return jsonify(load_data())


@app.route('/api/data', methods=['POST'])
def update_data():
    data = request.json
    save_data(data)
    return jsonify({'status': 'ok'})


# ── Saved projects ────────────────────────────────────────────
@app.route('/api/projects', methods=['GET'])
def list_projects():
    projects = []
    for fname in sorted(os.listdir(PROJECTS_DIR)):
        if fname.endswith('.json'):
            path = os.path.join(PROJECTS_DIR, fname)
            try:
                with open(path) as f:
                    d = json.load(f)
                meta = d.get('_meta', {})
                projects.append({
                    'id':         fname[:-5],
                    'name':       meta.get('name', fname[:-5]),
                    'customer':   d.get('project', {}).get('customer', ''),
                    'saved_at':   meta.get('saved_at', ''),
                })
            except Exception:
                pass
    return jsonify(projects)


@app.route('/api/projects', methods=['POST'])
def save_project():
    data = request.json
    name = data.get('_meta', {}).get('name') or \
           data.get('project', {}).get('customer') or 'Untitled'
    pid  = safe_filename(name) + '_' + datetime.now().strftime('%Y%m%d_%H%M%S')
    data['_meta'] = {
        'name':     name,
        'id':       pid,
        'saved_at': datetime.now().isoformat(timespec='seconds')
    }
    path = os.path.join(PROJECTS_DIR, pid + '.json')
    with open(path, 'w') as f:
        json.dump(data, f, indent=2)
    save_data(data)   # also update current working copy
    return jsonify({'status': 'ok', 'id': pid, 'name': name})


@app.route('/api/projects/<pid>', methods=['GET'])
def load_project(pid):
    path = os.path.join(PROJECTS_DIR, pid + '.json')
    if not os.path.exists(path):
        return jsonify({'error': 'Not found'}), 404
    with open(path) as f:
        data = json.load(f)
    # Always use the current global standard rate card + role catalog
    s = load_settings()
    if s.get('rate_card'):
        data['rate_card'] = s['rate_card']
    if s.get('role_catalog'):
        data['role_catalog'] = s['role_catalog']
    save_data(data)   # set as current working copy
    return jsonify(data)


@app.route('/api/settings', methods=['POST'])
def update_settings():
    s = request.json or {}
    existing = load_settings()
    if s.get('rate_card'):
        existing['rate_card'] = s['rate_card']
    if s.get('role_catalog'):
        existing['role_catalog'] = s['role_catalog']
    save_settings(existing)
    return jsonify({'status': 'ok'})


@app.route('/api/projects/<pid>', methods=['DELETE'])
def delete_project(pid):
    path = os.path.join(PROJECTS_DIR, pid + '.json')
    if os.path.exists(path):
        os.remove(path)
    return jsonify({'status': 'ok'})


@app.route('/api/export', methods=['POST'])
def export_excel():
    try:
        data = request.json
        if not data:
            return jsonify({'status': 'error', 'error': 'No data received'}), 400
        save_data(data)

        rate_map = {r['level']: r['rate'] for r in data['rate_card']}
        resources = data['resources']

        input_cost = sum(
            r['hours'] * rate_map.get(r['level'], 0)
            for r in resources
        )
        sell_cost = input_cost / 0.6 if input_cost > 0 else 0

        wb = build_workbook(data, input_cost, sell_cost, rate_map, resources)

        # Stream directly to browser — no disk write, no file lock conflicts
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        customer = data.get('project', {}).get('customer', 'PnL') or 'PnL'
        # sanitize for filesystem
        safe_customer = ''.join(c for c in customer if c.isalnum() or c in ' _-').strip().replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base = data.get('export_filename') or f"{safe_customer}_PnL"
        filename = f"{base}_{timestamp}.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        return jsonify({'status': 'error', 'error': str(e), 'detail': traceback.format_exc()}), 500


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def header_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def build_workbook(data, input_cost, sell_cost, rate_map, resources):
    wb = openpyxl.Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: PnL
    # -----------------------------------------------------------------------
    ws_pnl = wb.active
    ws_pnl.title = 'PnL'

    proj = data['project']
    att = data['attachments']
    fund = data['funding']
    appr = data['approvals']

    # Column widths
    col_widths = [22, 18, 22, 20, 18, 12, 14, 12, 12, 14, 14, 10, 14]
    for i, w in enumerate(col_widths, 1):
        ws_pnl.column_dimensions[get_column_letter(i)].width = w

    # Row 1: Company name
    ws_pnl.merge_cells('A1:M1')
    c = ws_pnl['A1']
    c.value = proj['company']
    c.font = Font(bold=True, size=16, color='FFFFFF')
    c.fill = header_fill('2D1B69')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_pnl.row_dimensions[1].height = 28

    # Row 2: Title
    ws_pnl.merge_cells('A2:M2')
    c = ws_pnl['A2']
    c.value = 'Project Profit and Loss Statement'
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.fill = header_fill('6D28D9')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_pnl.row_dimensions[2].height = 22

    # Row 3: Customer Name | value | ... | Date | value
    _set(ws_pnl, 3, 1, 'Customer Name', bold=True, fill='EDE9FE')
    ws_pnl.merge_cells('B3:J3')
    _set(ws_pnl, 3, 2, proj.get('customer', ''))
    _set(ws_pnl, 3, 11, 'Date', bold=True, fill='EDE9FE')
    ws_pnl.merge_cells('L3:M3')
    _set(ws_pnl, 3, 12, '')

    # Row 4: Location | value | ... | Reference | value
    _set(ws_pnl, 4, 1, 'Location', bold=True, fill='EDE9FE')
    ws_pnl.merge_cells('B4:J4')
    _set(ws_pnl, 4, 2, proj.get('location', ''))
    _set(ws_pnl, 4, 11, 'Reference', bold=True, fill='EDE9FE')
    ws_pnl.merge_cells('L4:M4')
    _set(ws_pnl, 4, 12, proj.get('reference', ''))

    # Row 5: Proposal Date | value | ... | Customer First Touch Point | value
    _set(ws_pnl, 5, 1, 'Proposal Date', bold=True, fill='EDE9FE')
    ws_pnl.merge_cells('B5:J5')
    proposal_date = proj.get('proposal_date') or datetime.today().strftime('%Y-%m-%d')
    _set(ws_pnl, 5, 2, proposal_date)
    ws_pnl.merge_cells('K5:M5')
    _set(ws_pnl, 5, 11, proj.get('customer_first_touch_point', ''))

    # Row 6: Proposal Value | sell_cost
    _set(ws_pnl, 6, 1, 'Proposal Value', bold=True, fill='EDE9FE')
    ws_pnl.merge_cells('B6:J6')
    _set(ws_pnl, 6, 2, round(sell_cost, 2), num_fmt='#,##0.00')

    ws_pnl.row_dimensions[6].height = 16

    # Row 7: Table headers
    headers_7 = [
        ('A7', 'Project Description'), ('C7', 'Delivery Method'),
        ('D7', 'Partner Details'), ('E7', 'Payment Terms'),
        ('F7', 'Duration\n(Months)'), ('G7', 'Revenue (USD)'),
        ('H7', 'Revenue Split'), ('I7', 'Cost Split (%)'),
        ('J7', 'Cost (USD)'), ('K7', 'Markup Value\n(USD)'),
        ('L7', 'Markup %'), ('M7', 'Gross Margin')
    ]
    ws_pnl.row_dimensions[7].height = 30
    for cell_ref, val in headers_7:
        c = ws_pnl[cell_ref]
        c.value = val
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.fill = header_fill('6D28D9')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws_pnl.merge_cells('A7:B7')

    # Resource rows in PnL (start row 8)
    pnl_roles = data.get('pnl_roles', [])
    start_row = 8
    markup = round(sell_cost - input_cost, 2)
    markup_pct = round(markup / input_cost, 4) if input_cost > 0 else 0
    gross_margin = round(markup / sell_cost, 4) if sell_cost > 0 else 0

    for idx, role in enumerate(pnl_roles):
        r = start_row + idx
        if idx == 0:
            _set(ws_pnl, r, 1, proj.get('project_description', proj['customer']),
                 bold=True, fill='F5F3FF')
        else:
            ws_pnl.merge_cells(f'A{r}:B{r}')
            _set(ws_pnl, r, 1, role.get('name', ''), fill='F2F2F2')
        ws_pnl.merge_cells(f'A{r}:B{r}')
        _set(ws_pnl, r, 3, role.get('name', ''))
        _set(ws_pnl, r, 4, role.get('partner', ''))
        _set(ws_pnl, r, 5, role.get('payment_terms', ''))
        _set(ws_pnl, r, 6, proj.get('duration_months', ''))
        if idx == 0:
            _set(ws_pnl, r, 7, round(sell_cost, 2), num_fmt='#,##0.00', fill='F0FDF4')
            _set(ws_pnl, r, 10, round(input_cost, 2), num_fmt='#,##0.00', fill='FEF3C7')
            _set(ws_pnl, r, 11, markup, num_fmt='#,##0.00', fill='F0FDF4')
            _set(ws_pnl, r, 12, markup_pct, num_fmt='0.0%')
            _set(ws_pnl, r, 13, gross_margin, num_fmt='0.0%')
        for col in range(1, 14):
            ws_pnl.cell(r, col).border = thin_border()

    end_roles_row = start_row + len(pnl_roles) - 1

    # Attachments section
    att_start = end_roles_row + 2
    _set(ws_pnl, att_start, 3, 'Yes/No', bold=True, fill='EDE9FE')
    _set(ws_pnl, att_start, 4, 'Reason, if not attached', bold=True, fill='EDE9FE')

    att_items = [
        ('Customer PO attached', att.get('customer_po', False)),
        ('Cloud4C Quote attached', att.get('cloud4c_quote', False)),
        ('Partner proposal attached', att.get('partner_proposal', False))
    ]
    for i, (label, val) in enumerate(att_items):
        r = att_start + 1 + i
        _set(ws_pnl, r, 1, label, bold=True)
        _set(ws_pnl, r, 3, 'Yes' if val else 'No')

    # Funding/Rebates section
    fund_start = att_start + len(att_items) + 2
    _set(ws_pnl, fund_start, 1, 'Funding / Rebates', bold=True, fill='6D28D9', color='FFFFFF')
    fund_headers = [('C', 'Method of recovery'), ('D', 'Reference'), ('E', 'Currency'), ('F', 'Value')]
    for col_letter, h in fund_headers:
        c = ws_pnl[f'{col_letter}{fund_start}']
        c.value = h
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = header_fill('6D28D9')
        c.border = thin_border()

    fund_items = [
        ('Marketing', fund.get('marketing', {})),
        ('Management', fund.get('management', {})),
        ('Discount', fund.get('discount', {}))
    ]
    for i, (label, fdata) in enumerate(fund_items):
        r = fund_start + 1 + i
        _set(ws_pnl, r, 1, label, bold=True)
        _set(ws_pnl, r, 3, fdata.get('method', ''))
        _set(ws_pnl, r, 4, fdata.get('reference', ''))
        _set(ws_pnl, r, 5, fdata.get('currency', 'USD'))
        v = fdata.get('value')
        _set(ws_pnl, r, 6, v if v is not None else '', num_fmt='#,##0.00' if v else None)

    total_row = fund_start + len(fund_items) + 1
    _set(ws_pnl, total_row, 5, 'Total', bold=True)
    fund_vals = [f.get('value') or 0 for _, f in fund_items]
    _set(ws_pnl, total_row, 6, sum(fund_vals), num_fmt='#,##0.00', bold=True, fill='FEF3C7')

    # Approvals
    appr_start = total_row + 2
    appr_items = [
        ('Prepared By', appr.get('prepared_by', '')),
        ('Reviewed By', appr.get('reviewed_by', '')),
        ('Approved By', appr.get('approved_by', ''))
    ]
    for i, (label, val) in enumerate(appr_items):
        r = appr_start + i
        _set(ws_pnl, r, 1, label, bold=True, fill='EDE9FE')
        ws_pnl.merge_cells(f'B{r}:G{r}')
        _set(ws_pnl, r, 2, val)

    # -----------------------------------------------------------------------
    # Sheet 2: Input Costing Calculation
    # -----------------------------------------------------------------------
    ws_cost = wb.create_sheet('Input Costing Calculation')

    cost_col_widths = [12, 12, 14, 8, 22, 6, 22, 6, 10, 6, 10, 10]
    for i, w in enumerate(cost_col_widths, 1):
        ws_cost.column_dimensions[get_column_letter(i)].width = w

    releases = data.get('releases', [])

    # Row 1: Release table header
    rel_headers = ['Release', 'Entities', '', 'EC+Time Off', 'Payroll Countries',
                   'EC Payroll', 'ABAP, CPI, PTP', 'BTP', 'RCM+ONB+OFFB', 'PMGM', 'LMS', 'Months']
    for col, h in enumerate(rel_headers, 1):
        c = ws_cost.cell(1, col)
        c.value = h
        if h:
            c.font = Font(bold=True, color='FFFFFF', size=9)
            c.fill = header_fill('2D1B69')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()

    # Rows 2..N: release data
    release_totals = {
        'entities': 0, 'ec_time_off': 0, 'payroll_countries': [],
        'ec_payroll': 0, 'abap_cpi_ptp': 0, 'btp': 0,
        'rcm_onb_offb': 0, 'pmgm': 0, 'lms': 0
    }
    for ri, rel in enumerate(releases):
        row = 2 + ri
        ws_cost.cell(row, 1).value = rel['name']
        ws_cost.cell(row, 2).value = rel['entities']
        ws_cost.cell(row, 4).value = rel.get('ec_time_off')
        ws_cost.cell(row, 5).value = rel.get('payroll_countries', '')
        ws_cost.cell(row, 6).value = rel.get('ec_payroll')
        ws_cost.cell(row, 7).value = rel.get('abap_cpi_ptp')
        ws_cost.cell(row, 8).value = rel.get('btp')
        ws_cost.cell(row, 9).value = rel.get('rcm_onb_offb')
        ws_cost.cell(row, 10).value = rel.get('pmgm')
        ws_cost.cell(row, 11).value = rel.get('lms')
        for col in range(1, 13):
            ws_cost.cell(row, col).border = thin_border()

        def add(key, val):
            if val:
                release_totals[key] += val
        add('entities', rel.get('entities', 0))
        add('ec_time_off', rel.get('ec_time_off') or 0)
        add('ec_payroll', rel.get('ec_payroll') or 0)
        add('abap_cpi_ptp', rel.get('abap_cpi_ptp') or 0)
        add('btp', rel.get('btp') or 0)
        add('rcm_onb_offb', rel.get('rcm_onb_offb') or 0)
        add('pmgm', rel.get('pmgm') or 0)
        add('lms', rel.get('lms') or 0)
        if rel.get('payroll_countries'):
            release_totals['payroll_countries'].append(rel['payroll_countries'])

    total_row_idx = 2 + len(releases)
    ws_cost.cell(total_row_idx, 1).value = 'Total'
    ws_cost.cell(total_row_idx, 2).value = release_totals['entities']
    ws_cost.cell(total_row_idx, 4).value = release_totals['ec_time_off']
    ws_cost.cell(total_row_idx, 5).value = ', '.join(release_totals['payroll_countries'])
    ws_cost.cell(total_row_idx, 6).value = release_totals['ec_payroll']
    ws_cost.cell(total_row_idx, 7).value = release_totals['abap_cpi_ptp']
    ws_cost.cell(total_row_idx, 8).value = release_totals['btp']
    ws_cost.cell(total_row_idx, 9).value = release_totals['rcm_onb_offb']
    ws_cost.cell(total_row_idx, 10).value = release_totals['pmgm']
    ws_cost.cell(total_row_idx, 11).value = release_totals['lms']
    ws_cost.cell(total_row_idx, 12).value = sum(
        (release_totals.get(k) or 0)
        for k in ['ec_time_off', 'ec_payroll', 'abap_cpi_ptp', 'btp', 'rcm_onb_offb', 'pmgm', 'lms']
    )
    for col in range(1, 13):
        c = ws_cost.cell(total_row_idx, col)
        c.font = Font(bold=True)
        c.fill = header_fill('EDE9FE')
        c.border = thin_border()

    # Blank row
    blank_row = total_row_idx + 1

    # Section header row
    sec_row = blank_row + 1
    _set_cost(ws_cost, sec_row, 1, 'HOURS', bold=True, fill='FEF9C3')
    _set_cost(ws_cost, sec_row, 3, 'Cost', bold=True, fill='DCFCE7')
    _set_cost(ws_cost, sec_row, 4, 'Level', bold=True, fill='EDE9FE')
    _set_cost(ws_cost, sec_row, 5, 'Role', bold=True, fill='EDE9FE')
    _set_cost(ws_cost, sec_row, 7, 'Role', bold=True, fill='F3E8FF')
    _set_cost(ws_cost, sec_row, 8, 'Level', bold=True, fill='F3E8FF')
    _set_cost(ws_cost, sec_row, 9, 'Hours', bold=True, fill='F3E8FF')
    _set_cost(ws_cost, sec_row, 11, 'Level', bold=True, fill='F0FDF4')
    _set_cost(ws_cost, sec_row, 12, 'Rate ($/hr)', bold=True, fill='F0FDF4')

    # Rate card (alongside resource rows)
    rate_card = data['rate_card']

    # Resource rows start
    res_start = sec_row + 1
    total_hours = 0
    total_cost_sum = 0

    for ri, res in enumerate(resources):
        r = res_start + ri
        rate = rate_map.get(res['level'], 0)
        cost = res['hours'] * rate
        total_hours += res['hours']
        total_cost_sum += cost

        _set_cost(ws_cost, r, 1, res['hours'])
        _set_cost(ws_cost, r, 3, round(cost, 2), num_fmt='#,##0.00')
        _set_cost(ws_cost, r, 4, res['level'])
        _set_cost(ws_cost, r, 5, res['role'])
        _set_cost(ws_cost, r, 7, res['role'])
        _set_cost(ws_cost, r, 8, res['level'])
        _set_cost(ws_cost, r, 9, res['hours'])

        # Rate card column
        if ri < len(rate_card):
            rc = rate_card[ri]
            _set_cost(ws_cost, r, 11, rc['level'], fill='F0FDF4')
            _set_cost(ws_cost, r, 12, rc['rate'], fill='F0FDF4', num_fmt='#,##0.00')

    # Fill remaining rate card rows if more rate levels than resources
    for ri in range(len(resources), len(rate_card)):
        r = res_start + ri
        rc = rate_card[ri]
        _set_cost(ws_cost, r, 11, rc['level'], fill='F0FDF4')
        _set_cost(ws_cost, r, 12, rc['rate'], fill='F0FDF4', num_fmt='#,##0.00')

    # Totals row
    max_rows = max(len(resources), len(rate_card))
    totals_row = res_start + max_rows
    _set_cost(ws_cost, totals_row, 1, total_hours, bold=True, fill='FEF9C3')
    _set_cost(ws_cost, totals_row, 3, round(total_cost_sum, 2), bold=True, fill='DCFCE7', num_fmt='#,##0.00')
    _set_cost(ws_cost, totals_row, 8, 'Total', bold=True)
    _set_cost(ws_cost, totals_row, 9, total_hours, bold=True, fill='FEF9C3')

    # Summary rows
    sum_row = totals_row + 2
    ws_cost.cell(sum_row, 2).value = 'proposal'
    _set_cost(ws_cost, sum_row, 3, 'Input Cost', bold=True, fill='FEF3C7')
    _set_cost(ws_cost, sum_row, 4, round(input_cost, 2), num_fmt='#,##0.00', fill='FEF3C7')

    _set_cost(ws_cost, sum_row + 1, 3, 'Sell Cost', bold=True, fill='F0FDF4')
    _set_cost(ws_cost, sum_row + 1, 4, round(sell_cost, 2), num_fmt='#,##0.00', fill='F0FDF4')

    return wb


def _set(ws, row, col, value, bold=False, fill=None, color='000000', num_fmt=None, align='left'):
    c = ws.cell(row, col)
    c.value = value
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(vertical='center', horizontal=align, wrap_text=True)
    if fill:
        c.fill = header_fill(fill)
    if num_fmt:
        c.number_format = num_fmt
    c.border = thin_border()
    return c


def _set_cost(ws, row, col, value, bold=False, fill=None, num_fmt=None):
    c = ws.cell(row, col)
    c.value = value
    c.font = Font(bold=bold)
    c.alignment = Alignment(vertical='center')
    if fill:
        c.fill = header_fill(fill)
    if num_fmt:
        c.number_format = num_fmt
    c.border = thin_border()
    return c


if __name__ == '__main__':
    app.run(debug=True, port=5000)
